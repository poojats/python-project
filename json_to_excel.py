import json
import openpyxl

# Replace 'Your_JSON_File.json' with the actual path to your JSON file
json_file_path = r"C:Example.json"

# Read JSON data from file and clean invalid control characters
with open(json_file_path, 'r', encoding='utf-8') as file:
    json_data_str = file.read()

# Remove invalid control characters
cleaned_json_data_str = ''.join(char for char in json_data_str if char.isprintable())

# Load the cleaned JSON data into a dictionary
json_data = json.loads(cleaned_json_data_str)

# Create Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Sheet 1"  # Rename the first sheet
sheet1 = workbook.create_sheet("Sheet 2")  # Create the second sheet

# Write headers for the first sheet
sheet.cell(row=1, column=1, value='Item No')
sheet.cell(row=1, column=2, value='DEFAULT_SKU')
sheet.cell(row=1, column=3, value='ATG_DEFAULT_CATEGORY')
sheet.cell(row=1, column=4, value='DESCRIPTION_PRODUCT')
sheet.cell(row=1, column=5, value='DESCRIPTION_PRODUCT_VN')
sheet.cell(row=1, column=6, value='DESCRIPTION_PRODUCT_SHORT')
sheet.cell(row=1, column=7, value='DESCRIPTION_PRODUCT_SHORT_VN')
sheet.cell(row=1, column=8, value='PRODUCT_TYPE')
sheet.cell(row=1, column=9, value='MATERIAL')
sheet.cell(row=1, column=10, value='WEB_INCLUDED')
sheet.cell(row=1, column=11, value='INSTALLATION_TYPE')
sheet.cell(row=1, column=12, value='COLLECTION')
sheet.cell(row=1, column=13, value='BRAND_NAME')

# Write headers for the second sheet
sheet1.cell(row=1, column=1, value='Item No')
sheet1.cell(row=1, column=2, value='WEB_INCLUDED')
sheet1.cell(row=1, column=3, value='COLLECTION')
sheet1.cell(row=1, column=4, value='BRAND_NAME')
sheet1.cell(row=1, column=5, value='SKUs')  # Modified to store all SKUs in one cell
sheet1.cell(row=1, column=6, value='LIST_PRICES')  # List prices for all SKUs in one cell
sheet1.cell(row=1, column=7, value='COLOR_FINISH_CODES')  # Color finish codes for all SKUs in one cell
sheet1.cell(row=1, column=8, value='COLOR_FINISH_NAMES')  # Color finish names for all SKUs in one cell
sheet1.cell(row=1, column=9, value='IMG_SWATCHES')  # Image swatches for all SKUs in one cell
sheet1.cell(row=1, column=10, value='IMG_ITEM_ISOS')  # Image item ISOs for all SKUs in one cell
sheet1.cell(row=1, column=11, value='WEB_INCLUDED_SKUs')  # Web included status for all SKUs in one cell
sheet1.cell(row=1, column=12, value='RETAILER_1_IMGs')  # Retailer Url for all SKUs in one cell
sheet1.cell(row=1, column=13, value='RETAILER_1_URLs')  # Retailer Img for all SKUs in one cell
sheet1.cell(row=1, column=14, value='RETAILER_2_IMGs')  # Retailer Url for all SKUs in one cell
sheet1.cell(row=1, column=15, value='RETAILER_2_URLs')  # Retailer Img for all SKUs in one cell

# Populate Excel with data
for item_id, item_data in json_data.items():
    item_no = item_data.get('itemNo')
    default_sku = item_data.get('attributes', {}).get('DEFAULT_SKU')
    default_category = item_data.get('attributes', {}).get('ATG_DEFAULT_CATEGORY')
    description_product = item_data.get('attributes', {}).get('DESCRIPTION_PRODUCT')
    description_product_VN = item_data.get('attributes', {}).get('DESCRIPTION_PRODUCT_VN')
    description_product_short = item_data.get('attributes', {}).get('DESCRIPTION_PRODUCT_SHORT')
    description_product_short_VN = item_data.get('attributes', {}).get('DESCRIPTION_PRODUCT_SHORT_VN')

    product_type = item_data.get('attributes', {}).get('PRODUCT_TYPE')
    material = item_data.get('attributes', {}).get('MATERIAL')
    web_included = item_data.get('attributes', {}).get('WEB_INCLUDED')

    installation_data = item_data.get('attributes', {}).get('INSTALLATION_TYPE', {})
    installation_type = next(iter(installation_data.values()), None)

    collection_data = item_data.get('attributes', {}).get('COLLECTION', {})
    collection = next(iter(collection_data.values()), None)

    brand_data = item_data.get('attributes', {}).get('BRAND_NAME', {})
    brand = next(iter(brand_data.values()), None)

    # Extract SKU data
    skus = item_data.get('skus', [])
    sku_list = []
    list_price_list = []
    color_finish_code_list = []
    color_finish_name_list = []
    img_swatch_list = []
    img_item_iso_list = []
    web_included_sku_list = []

    for sku_data in skus:
        sku = sku_data.get('sku')
        sku_attributes = sku_data.get('skuAttributes', {})
        list_price = sku_attributes.get('LIST_PRICE')
        # id = sku_attributes.get()
        color_finish_code = sku_attributes.get('COLOR_FINISH_CODE')
        color_finish_name = sku_attributes.get('COLOR_FINISH_NAME')
        img_swatch = sku_attributes.get('IMG_SWATCH')
        img_item_iso = sku_attributes.get('IMG_ITEM_ISO')
        web_included_sku = sku_attributes.get('WEB_INCLUDED')
        retailer_1_url = sku_attributes.get('RETAILER_1_URL')
        retailer_1_img = sku_attributes.get('RETAILER_1_IMG')
        retailer_2_url = sku_attributes.get('RETAILER_2_URL')
        retailer_2_img = sku_attributes.get('RETAILER_2_IMG')

        sku_list.append(sku)
        # list_price_list.append(list_price)
        # color_finish_code_list.append(color_finish_code)
        # color_finish_name_list.append(color_finish_name)
        # img_swatch_list.append(img_swatch)
        # img_item_iso_list.append(img_item_iso)
        # web_included_sku_list.append(web_included_sku)

    # Append data to the first sheet
    sheet.append([item_no, default_sku, default_category, description_product, description_product_VN, description_product_short, description_product_short_VN, product_type, material, web_included, installation_type, collection, brand])

    # Append all SKUs and their attributes to the second sheet, joined by commas
    sheet1.append([
        item_no,
        web_included,
        collection,
        brand,
        ', '.join(sku_list),
        list_price,
        color_finish_code,
        color_finish_name,
        img_swatch,
        img_item_iso,
        web_included_sku,
        retailer_1_url,
        retailer_1_img,
        retailer_2_url,
        retailer_2_img
    ])

# Save Excel file
excel_file_path = r'C:example.xlsx'
workbook.save(excel_file_path)

print(f"Successfully created and saved Excel file: {excel_file_path}")
